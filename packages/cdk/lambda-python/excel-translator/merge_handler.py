"""
Excel Translation - Merge Handler for Step Functions

This Lambda:
1. Loads all translation results from S3
2. Loads original Excel file
3. Applies translations to all cells
4. Uploads translated file to S3
5. Updates job status to COMPLETED
"""

import json
import os
import uuid
from datetime import datetime

import boto3
from aws_lambda_powertools import Logger
from aws_lambda_powertools.utilities.typing import LambdaContext
from openpyxl import load_workbook

# For .xls support
import xlrd
from openpyxl import Workbook

logger = Logger()

s3_client = boto3.client("s3")
dynamodb_client = boto3.client("dynamodb")

BUCKET_NAME = os.environ.get("BUCKET_NAME", "")
JOB_TABLE_NAME = os.environ.get("JOB_TABLE_NAME", "")


def update_job_status(job_id: str, status: str, **kwargs) -> None:
    """Update job status in DynamoDB"""
    if not JOB_TABLE_NAME or not job_id:
        return

    update_expr = "SET #status = :status"
    expr_names = {"#status": "status"}
    expr_values = {":status": {"S": status}}

    for key, value in kwargs.items():
        update_expr += f", {key} = :{key}"
        if isinstance(value, dict):
            expr_values[f":{key}"] = {"S": json.dumps(value)}
        else:
            expr_values[f":{key}"] = {"S": str(value)}

    try:
        dynamodb_client.update_item(
            TableName=JOB_TABLE_NAME,
            Key={"jobId": {"S": job_id}},
            UpdateExpression=update_expr,
            ExpressionAttributeNames=expr_names,
            ExpressionAttributeValues=expr_values,
        )
    except Exception as e:
        logger.warning(f"Failed to update job status: {e}")


def convert_xls_to_xlsx(xls_path: str, xlsx_path: str) -> None:
    """Convert .xls file to .xlsx format."""
    xls_book = xlrd.open_workbook(xls_path, formatting_info=False)
    xlsx_book = Workbook()

    if "Sheet" in xlsx_book.sheetnames:
        del xlsx_book["Sheet"]

    for sheet_idx in range(xls_book.nsheets):
        xls_sheet = xls_book.sheet_by_index(sheet_idx)
        xlsx_sheet = xlsx_book.create_sheet(title=xls_sheet.name)

        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                cell = xls_sheet.cell(row_idx, col_idx)
                value = cell.value

                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate.xldate_as_datetime(cell.value, xls_book.datemode)
                    except Exception:
                        pass
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    value = bool(cell.value)
                elif cell.ctype == xlrd.XL_CELL_ERROR:
                    value = None

                xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=value)

    xlsx_book.save(xlsx_path)


def is_xls_file(filename: str) -> bool:
    """Check if file is .xls format (not .xlsx)"""
    return filename.lower().endswith(".xls") and not filename.lower().endswith(".xlsx")


def cleanup_work_data(job_id: str) -> None:
    """Clean up temporary work data from S3"""
    work_prefix = f"excel-work/{job_id}/"
    try:
        response = s3_client.list_objects_v2(Bucket=BUCKET_NAME, Prefix=work_prefix)
        if "Contents" in response:
            objects = [{"Key": obj["Key"]} for obj in response["Contents"]]
            if objects:
                s3_client.delete_objects(Bucket=BUCKET_NAME, Delete={"Objects": objects})
                logger.info(f"Cleaned up {len(objects)} work files")
    except Exception as e:
        logger.warning(f"Failed to cleanup work data: {e}")


@logger.inject_lambda_context
def lambda_handler(event: dict, context: LambdaContext) -> dict:
    """
    Merge handler for Step Functions.

    Input:
    {
        "jobId": "uuid",
        "s3Key": "uploads/xxx/file.xlsx",
        "workDataKey": "excel-work/jobId/work_data.json",
        "translationResults": [
            {"batchId": 0, "translationKey": "...", "translatedCount": 100, "success": true},
            ...
        ],
        "stats": {...}
    }

    Output:
    {
        "jobId": "uuid",
        "outputS3Key": "translated/xxx/file_translated.xlsx",
        "downloadUrl": "presigned-url",
        "stats": {...}
    }
    """
    job_id = event.get("jobId")
    s3_key = event.get("s3Key")
    work_data_key = event.get("workDataKey")
    translation_results = event.get("translationResults", [])
    stats = event.get("stats", {})

    logger.info(f"Merging translations for job {job_id}")

    update_job_status(
        job_id,
        "MERGING",
        progress=json.dumps({
            "phase": "merging",
            "percent": 90
        })
    )

    # Load work data
    response = s3_client.get_object(Bucket=BUCKET_NAME, Key=work_data_key)
    work_data = json.loads(response["Body"].read().decode("utf-8"))
    cells = work_data.get("cells", [])

    logger.info(f"Loaded work data with {len(cells)} cells to process")

    # Load all translations
    all_translations = {}
    for result in translation_results:
        if result.get("success"):
            translation_key = result.get("translationKey")
            try:
                response = s3_client.get_object(Bucket=BUCKET_NAME, Key=translation_key)
                data = json.loads(response["Body"].read().decode("utf-8"))
                all_translations.update(data.get("translations", {}))
            except Exception as e:
                logger.warning(f"Failed to load translation {translation_key}: {e}")

    logger.info(f"Loaded {len(all_translations)} translations from {len(translation_results)} batches")

    # Download original file
    filename = os.path.basename(s3_key)
    is_xls = is_xls_file(filename)
    unique_id = uuid.uuid4()

    if is_xls:
        tmp_download = f"/tmp/input_{unique_id}.xls"
        tmp_input = f"/tmp/input_{unique_id}_converted.xlsx"
    else:
        tmp_download = f"/tmp/input_{unique_id}.xlsx"
        tmp_input = tmp_download

    tmp_output = f"/tmp/output_{unique_id}.xlsx"

    s3_client.download_file(BUCKET_NAME, s3_key, tmp_download)

    if is_xls:
        convert_xls_to_xlsx(tmp_download, tmp_input)

    # Load workbook and apply translations
    wb = load_workbook(tmp_input)

    translated_count = 0
    for cell_info in cells:
        sheet_name = cell_info.get("sheet")
        coord = cell_info.get("coord")
        original_text = cell_info.get("text")

        if original_text in all_translations:
            try:
                wb[sheet_name][coord].value = all_translations[original_text]
                translated_count += 1
            except Exception as e:
                logger.warning(f"Failed to apply translation to {sheet_name}!{coord}: {e}")

    # Save translated workbook
    wb.save(tmp_output)
    logger.info(f"Applied {translated_count} translations to workbook")

    # Upload to S3
    name, _ = os.path.splitext(filename)
    output_filename = f"{name}_translated.xlsx"
    output_s3_key = f"translated/{uuid.uuid4()}/{output_filename}"

    s3_client.upload_file(tmp_output, BUCKET_NAME, output_s3_key)
    logger.info(f"Uploaded translated file to S3: {output_s3_key}")

    # Generate presigned URL
    presigned_url = s3_client.generate_presigned_url(
        "get_object",
        Params={"Bucket": BUCKET_NAME, "Key": output_s3_key},
        ExpiresIn=3600,
    )

    # Clean up temp files
    for f in [tmp_download, tmp_input, tmp_output]:
        if os.path.exists(f):
            os.remove(f)

    # Clean up S3 work data
    cleanup_work_data(job_id)

    # Update final stats
    stats["translatedCells"] = translated_count
    stats["sheetsProcessed"] = len(wb.worksheets)

    # Update job status to COMPLETED
    update_job_status(
        job_id,
        "COMPLETED",
        outputS3Key=output_s3_key,
        downloadUrl=presigned_url,
        stats=stats,
        completedAt=datetime.utcnow().isoformat(),
    )

    logger.info(f"Translation job {job_id} completed successfully")

    return {
        "jobId": job_id,
        "outputS3Key": output_s3_key,
        "downloadUrl": presigned_url,
        "stats": stats
    }

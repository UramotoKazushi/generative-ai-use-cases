"""
Excel Translator Lambda Handler

Translates Excel files while preserving all formatting (fonts, borders, colors, etc.)
Supports async processing with DynamoDB job status tracking.
Features:
- Translation caching for duplicate text optimization
- Smart cell filtering (skip numbers, URLs, emails, dates)
- Support for both .xlsx and .xls formats
"""

import json
import os
import re
import time
import uuid
from datetime import datetime
from typing import Any

import boto3
from botocore.exceptions import ClientError
from aws_lambda_powertools import Logger
from aws_lambda_powertools.utilities.typing import LambdaContext
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

# For .xls support
import xlrd
from openpyxl import Workbook

logger = Logger()

# Translation cache to avoid duplicate API calls
translation_cache: dict[str, str] = {}


def convert_xls_to_xlsx(xls_path: str, xlsx_path: str) -> None:
    """
    Convert .xls file to .xlsx format.
    Note: Some advanced formatting may not be preserved.
    """
    xls_book = xlrd.open_workbook(xls_path, formatting_info=False)
    xlsx_book = Workbook()

    # Remove default sheet
    if "Sheet" in xlsx_book.sheetnames:
        del xlsx_book["Sheet"]

    for sheet_idx in range(xls_book.nsheets):
        xls_sheet = xls_book.sheet_by_index(sheet_idx)
        xlsx_sheet = xlsx_book.create_sheet(title=xls_sheet.name)

        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                cell = xls_sheet.cell(row_idx, col_idx)
                value = cell.value

                # Handle different cell types
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate.xldate_as_datetime(cell.value, xls_book.datemode)
                    except Exception:
                        pass
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    value = bool(cell.value)
                elif cell.ctype == xlrd.XL_CELL_ERROR:
                    value = None

                xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=value)

    xlsx_book.save(xlsx_path)
    logger.info(f"Converted .xls to .xlsx: {xls_path} -> {xlsx_path}")


def is_xls_file(filename: str) -> bool:
    """Check if file is .xls format (not .xlsx)"""
    return filename.lower().endswith(".xls") and not filename.lower().endswith(".xlsx")


s3_client = boto3.client("s3")
dynamodb_client = boto3.client("dynamodb")
bedrock_client = boto3.client("bedrock-runtime", region_name=os.environ.get("MODEL_REGION", "us-east-1"))

BUCKET_NAME = os.environ.get("BUCKET_NAME", "")
MODEL_ID = os.environ.get("MODEL_ID", "us.anthropic.claude-3-5-haiku-20241022-v1:0")
JOB_TABLE_NAME = os.environ.get("JOB_TABLE_NAME", "")


def update_job_status(job_id: str, status: str, **kwargs) -> None:
    """Update job status in DynamoDB"""
    if not JOB_TABLE_NAME or not job_id:
        return

    update_expr = "SET #status = :status"
    expr_names = {"#status": "status"}
    expr_values = {":status": {"S": status}}

    for key, value in kwargs.items():
        update_expr += f", {key} = :{key}"
        if isinstance(value, dict):
            expr_values[f":{key}"] = {"S": json.dumps(value)}
        else:
            expr_values[f":{key}"] = {"S": str(value)}

    try:
        dynamodb_client.update_item(
            TableName=JOB_TABLE_NAME,
            Key={"jobId": {"S": job_id}},
            UpdateExpression=update_expr,
            ExpressionAttributeNames=expr_names,
            ExpressionAttributeValues=expr_values,
        )
    except Exception as e:
        logger.warning(f"Failed to update job status: {e}")


def bulk_translate_unique_texts(
    unique_texts: list[str],
    source_lang: str,
    target_lang: str,
    job_id: str = None,
    total_cells: int = 0,
) -> dict[str, str]:
    """
    Translate unique texts in large batches using JSON format.
    Returns a dict mapping original text to translated text.

    This is much more efficient than translating cell by cell:
    - 687 cells with 200 unique texts = 2-3 API calls instead of 28
    """
    if not unique_texts:
        return {}

    translations = {}
    texts_to_translate = []

    # Check cache first
    for text in unique_texts:
        cache_key = f"{source_lang}:{target_lang}:{text}"
        if cache_key in translation_cache:
            translations[text] = translation_cache[cache_key]
        else:
            texts_to_translate.append(text)

    if translations:
        logger.info(f"Cache hits: {len(translations)}/{len(unique_texts)} unique texts")

    if not texts_to_translate:
        return translations

    # Large batch size for efficiency - Claude can handle 100+ texts easily
    BATCH_SIZE = 100
    max_retries = 5
    total_batches = (len(texts_to_translate) + BATCH_SIZE - 1) // BATCH_SIZE

    logger.info(f"Translating {len(texts_to_translate)} unique texts in {total_batches} batches")

    for batch_idx in range(0, len(texts_to_translate), BATCH_SIZE):
        batch = texts_to_translate[batch_idx : batch_idx + BATCH_SIZE]
        batch_num = batch_idx // BATCH_SIZE + 1

        # Create JSON input for structured translation
        input_data = [{"id": i, "text": text} for i, text in enumerate(batch)]

        prompt = f"""Translate the following {source_lang} texts to {target_lang}.

IMPORTANT RULES:
- Return ONLY a valid JSON array with translations
- Each item must have "id" (same as input) and "translation" fields
- Preserve numbers, special characters, and formatting
- If text contains only numbers/symbols, return as-is

Input:
{json.dumps(input_data, ensure_ascii=False)}

Output format (JSON array only, no other text):
[{{"id": 0, "translation": "..."}}, {{"id": 1, "translation": "..."}}, ...]"""

        success = False
        for attempt in range(max_retries):
            try:
                response = bedrock_client.converse(
                    modelId=MODEL_ID,
                    messages=[{"role": "user", "content": [{"text": prompt}]}],
                    inferenceConfig={"maxTokens": 16384, "temperature": 0.1},
                )
                result = response["output"]["message"]["content"][0]["text"].strip()

                # Parse JSON response
                # Handle potential markdown code blocks
                if result.startswith("```"):
                    result = result.split("```")[1]
                    if result.startswith("json"):
                        result = result[4:]
                    result = result.strip()

                try:
                    parsed = json.loads(result)
                    for item in parsed:
                        idx = item.get("id")
                        translated = item.get("translation", "")
                        if idx is not None and 0 <= idx < len(batch):
                            original_text = batch[idx]
                            translations[original_text] = translated
                            # Store in cache
                            cache_key = f"{source_lang}:{target_lang}:{original_text}"
                            translation_cache[cache_key] = translated
                except json.JSONDecodeError:
                    # Fallback: try to parse line by line if JSON fails
                    logger.warning("JSON parse failed, attempting line-by-line parse")
                    for line in result.split("\n"):
                        line = line.strip()
                        if '"id"' in line and '"translation"' in line:
                            try:
                                item = json.loads(line.rstrip(","))
                                idx = item.get("id")
                                translated = item.get("translation", "")
                                if idx is not None and 0 <= idx < len(batch):
                                    original_text = batch[idx]
                                    translations[original_text] = translated
                                    cache_key = f"{source_lang}:{target_lang}:{original_text}"
                                    translation_cache[cache_key] = translated
                            except:
                                continue

                success = True
                logger.info(f"Translated batch {batch_num}/{total_batches} ({len(batch)} texts)")
                break

            except ClientError as e:
                if e.response["Error"]["Code"] == "ThrottlingException":
                    if attempt < max_retries - 1:
                        wait_time = (2 ** (attempt + 1)) + (time.time() % 1)
                        logger.info(f"Throttled, waiting {wait_time:.1f}s before retry {attempt + 2}/{max_retries}")
                        time.sleep(wait_time)
                        continue
                logger.warning(f"Batch translation failed: {e}")
                break
            except Exception as e:
                logger.warning(f"Batch translation failed: {e}")
                break

        # Fallback for any missing translations in this batch
        if not success or any(text not in translations for text in batch):
            missing = [t for t in batch if t not in translations]
            if missing:
                logger.info(f"Retrying {len(missing)} missing translations individually")
                for text in missing:
                    translated = translate_single_text(text, source_lang, target_lang)
                    translations[text] = translated

        # Update progress
        if job_id and total_cells > 0:
            translated_count = len(translations)
            update_job_status(
                job_id,
                "PROCESSING",
                progress=json.dumps({
                    "phase": "translating_unique_texts",
                    "unique_translated": translated_count,
                    "unique_total": len(unique_texts),
                    "batch_progress": f"{batch_num}/{total_batches}",
                    "percent": int(translated_count / len(unique_texts) * 100),
                }),
            )

        # Small delay between batches to avoid throttling
        if batch_idx + BATCH_SIZE < len(texts_to_translate):
            time.sleep(0.5)

    return translations


def translate_single_text(text: str, source_lang: str, target_lang: str, max_retries: int = 3) -> str:
    """Translate a single text as fallback"""
    if not text or not text.strip():
        return text

    cache_key = f"{source_lang}:{target_lang}:{text}"
    if cache_key in translation_cache:
        return translation_cache[cache_key]

    prompt = f"""Translate to {target_lang}. Output only the translation:
{text}"""

    for attempt in range(max_retries):
        try:
            response = bedrock_client.converse(
                modelId=MODEL_ID,
                messages=[{"role": "user", "content": [{"text": prompt}]}],
                inferenceConfig={"maxTokens": 1024, "temperature": 0.1},
            )
            translated = response["output"]["message"]["content"][0]["text"].strip()
            translation_cache[cache_key] = translated
            return translated
        except ClientError as e:
            if e.response["Error"]["Code"] == "ThrottlingException" and attempt < max_retries - 1:
                time.sleep(2 ** (attempt + 1))
                continue
            return text
        except:
            return text

    return text


def should_skip_text(text: str) -> bool:
    """
    Check if text should be skipped from translation.
    Returns True if the text doesn't need translation.
    Enhanced filtering to reduce unnecessary API calls.
    """
    text = text.strip()

    # Empty or whitespace only
    if not text:
        return True

    # Numbers only (including decimals, negatives, percentages, with spaces)
    if re.match(r'^-?[\d,\s]+\.?\d*%?$', text):
        return True

    # Date patterns (various formats)
    date_patterns = [
        r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}$',  # 2024-01-15, 2024/01/15
        r'^\d{1,2}[-/]\d{1,2}[-/]\d{4}$',  # 01-15-2024, 01/15/2024
        r'^\d{1,2}[-/]\d{1,2}[-/]\d{2}$',  # 01-15-24, 01/15/24
        r'^\d{4}年\d{1,2}月\d{1,2}日$',     # 2024年1月15日
    ]
    for pattern in date_patterns:
        if re.match(pattern, text):
            return True

    # Time patterns
    if re.match(r'^\d{1,2}:\d{2}(:\d{2})?(\s*[APap][Mm])?$', text):
        return True

    # URLs
    if re.match(r'^https?://', text, re.IGNORECASE):
        return True

    # Email addresses
    if re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', text):
        return True

    # Phone numbers (various formats)
    if re.match(r'^[\d\s\-\+\(\)]+$', text) and len(re.sub(r'\D', '', text)) >= 7:
        return True

    # Symbols and punctuation only (no letters)
    if not any(c.isalpha() for c in text):
        return True

    # Currency values
    if re.match(r'^[$€£¥₩]\s*[\d,]+\.?\d*$', text):
        return True
    if re.match(r'^[\d,]+\.?\d*\s*[$€£¥₩円]$', text):
        return True

    # English/ASCII only text (no need to translate if already in target language or code/identifiers)
    # Skip if text contains only ASCII letters, numbers, and common punctuation
    if re.match(r'^[A-Za-z0-9\s\.,;:!?\'"()\-_@#$%&*+=/<>\\|{}\[\]`~]+$', text):
        # But don't skip if it's a meaningful English sentence (has spaces and multiple words)
        words = text.split()
        if len(words) <= 2:
            # Short English text like "OK", "Yes", "ID", "No." - skip
            return True
        # Longer English text might need translation depending on context
        # Skip common technical terms and identifiers
        if re.match(r'^[A-Z][A-Za-z0-9_]+$', text):  # CamelCase identifiers
            return True
        if re.match(r'^[a-z_][a-z0-9_]*$', text):  # snake_case identifiers
            return True

    # File paths and technical identifiers
    if re.match(r'^[A-Za-z]:\\', text) or text.startswith('/'):
        return True

    return False


def is_translatable_cell(cell: Cell) -> bool:
    """Check if a cell contains translatable text"""
    if cell.value is None:
        return False
    if not isinstance(cell.value, str):
        return False
    if not cell.value.strip():
        return False
    # Skip cells that are formulas
    if str(cell.value).startswith("="):
        return False
    # Skip cells that don't need translation
    if should_skip_text(cell.value):
        return False
    return True


def translate_excel(input_path: str, output_path: str, source_lang: str, target_lang: str, job_id: str = None) -> dict:
    """
    Translate an Excel file while preserving all formatting.

    Uses unique text aggregation for efficiency:
    - Extract unique texts from ALL sheets
    - Translate unique texts in large batches (100 per API call)
    - Apply translations back to all cells

    This reduces API calls dramatically (e.g., 687 cells with 200 unique texts = 2-3 API calls)
    """
    # Load workbook preserving all formatting
    wb = load_workbook(input_path)

    stats = {"total_cells": 0, "translated_cells": 0, "sheets_processed": 0, "total_sheets": len(wb.worksheets)}

    # Phase 1: Collect ALL translatable cells and their texts from ALL sheets
    logger.info("Phase 1: Collecting texts from all sheets...")
    all_cells_info: list[tuple[str, str, str]] = []  # (sheet_name, coord, text)
    unique_texts: set[str] = set()
    total_cell_count = 0

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                total_cell_count += 1
                if is_translatable_cell(cell):
                    text = cell.value
                    all_cells_info.append((sheet.title, cell.coordinate, text))
                    unique_texts.add(text)

    stats["total_cells"] = total_cell_count
    stats["total_translatable"] = len(all_cells_info)
    stats["unique_texts"] = len(unique_texts)

    logger.info(f"Found {len(all_cells_info)} translatable cells with {len(unique_texts)} unique texts")

    if job_id:
        update_job_status(
            job_id,
            "PROCESSING",
            progress=json.dumps({
                "phase": "collecting_texts",
                "total_translatable": len(all_cells_info),
                "unique_texts": len(unique_texts),
                "percent": 5,
            }),
        )

    if not unique_texts:
        logger.info("No translatable text found")
        wb.save(output_path)
        return stats

    # Phase 2: Translate all unique texts at once
    logger.info("Phase 2: Translating unique texts in batches...")
    if job_id:
        update_job_status(
            job_id,
            "PROCESSING",
            progress=json.dumps({
                "phase": "translating",
                "unique_texts": len(unique_texts),
                "percent": 10,
            }),
        )

    translations = bulk_translate_unique_texts(
        list(unique_texts),
        source_lang,
        target_lang,
        job_id,
        len(all_cells_info),
    )

    logger.info(f"Translated {len(translations)} unique texts")

    # Phase 3: Apply translations back to all cells
    logger.info("Phase 3: Applying translations to cells...")
    if job_id:
        update_job_status(
            job_id,
            "PROCESSING",
            progress=json.dumps({
                "phase": "applying_translations",
                "percent": 90,
            }),
        )

    translated_count = 0
    for sheet_name, coord, original_text in all_cells_info:
        sheet = wb[sheet_name]
        if original_text in translations:
            sheet[coord].value = translations[original_text]
            translated_count += 1
        else:
            # Fallback: translate individually if not in cache
            translated = translate_single_text(original_text, source_lang, target_lang)
            sheet[coord].value = translated
            translated_count += 1

    stats["translated_cells"] = translated_count
    stats["sheets_processed"] = len(wb.worksheets)

    # Save the translated workbook
    wb.save(output_path)
    logger.info(f"Saved translated workbook with {translated_count} translated cells")

    return stats


@logger.inject_lambda_context
def lambda_handler(event: dict, context: LambdaContext) -> dict:
    """
    Lambda handler for Excel translation (async mode).

    Expected event format (async invocation from startExcelTranslation):
    {
        "jobId": "uuid",
        "s3Key": "uploads/xxx/file.xlsx",
        "sourceLanguage": "Japanese",
        "targetLanguage": "English"
    }

    Updates DynamoDB with job status as it processes.
    """
    job_id = event.get("jobId")
    s3_key = event.get("s3Key")
    source_lang = event.get("sourceLanguage", "Japanese")
    target_lang = event.get("targetLanguage", "English")

    # For backward compatibility, also check body
    if not s3_key:
        body = event.get("body")
        if isinstance(body, str):
            body = json.loads(body)
        elif body:
            s3_key = body.get("s3Key")
            source_lang = body.get("sourceLanguage", "Japanese")
            target_lang = body.get("targetLanguage", "English")

    try:
        if not s3_key:
            error_msg = "s3Key is required"
            if job_id:
                update_job_status(job_id, "FAILED", error=error_msg, failedAt=datetime.utcnow().isoformat())
            return {"statusCode": 400, "body": json.dumps({"error": error_msg})}

        if not BUCKET_NAME:
            error_msg = "BUCKET_NAME not configured"
            if job_id:
                update_job_status(job_id, "FAILED", error=error_msg, failedAt=datetime.utcnow().isoformat())
            return {"statusCode": 500, "body": json.dumps({"error": error_msg})}

        # Update status to PROCESSING
        if job_id:
            update_job_status(job_id, "PROCESSING", startedAt=datetime.utcnow().isoformat())

        logger.info(f"Processing file: {s3_key}, {source_lang} -> {target_lang}, jobId: {job_id}")

        # Determine file type
        filename = os.path.basename(s3_key)
        is_xls = is_xls_file(filename)

        # Create temp file paths
        unique_id = uuid.uuid4()
        if is_xls:
            tmp_download = f"/tmp/input_{unique_id}.xls"
            tmp_input = f"/tmp/input_{unique_id}_converted.xlsx"
        else:
            tmp_download = f"/tmp/input_{unique_id}.xlsx"
            tmp_input = tmp_download
        tmp_output = f"/tmp/output_{unique_id}.xlsx"

        # Download file from S3
        s3_client.download_file(BUCKET_NAME, s3_key, tmp_download)
        logger.info(f"Downloaded file from S3: {s3_key}")

        # Convert .xls to .xlsx if necessary
        if is_xls:
            convert_xls_to_xlsx(tmp_download, tmp_input)
            logger.info(f"Converted .xls to .xlsx for processing")

        # Translate the Excel file
        stats = translate_excel(tmp_input, tmp_output, source_lang, target_lang, job_id)
        logger.info(f"Translation complete: {stats}")

        # Generate output S3 key (always output as .xlsx for compatibility)
        name, _ = os.path.splitext(filename)
        output_filename = f"{name}_translated.xlsx"
        output_s3_key = f"translated/{uuid.uuid4()}/{output_filename}"

        # Upload translated file to S3
        s3_client.upload_file(tmp_output, BUCKET_NAME, output_s3_key)
        logger.info(f"Uploaded translated file to S3: {output_s3_key}")

        # Clean up temp files
        if is_xls and os.path.exists(tmp_download):
            os.remove(tmp_download)
        if os.path.exists(tmp_input):
            os.remove(tmp_input)
        if os.path.exists(tmp_output):
            os.remove(tmp_output)

        # Generate presigned URL for download
        presigned_url = s3_client.generate_presigned_url(
            "get_object",
            Params={"Bucket": BUCKET_NAME, "Key": output_s3_key},
            ExpiresIn=3600,  # 1 hour
        )

        # Update job status to COMPLETED
        if job_id:
            update_job_status(
                job_id,
                "COMPLETED",
                outputS3Key=output_s3_key,
                downloadUrl=presigned_url,
                stats=stats,
                completedAt=datetime.utcnow().isoformat(),
            )

        return {
            "statusCode": 200,
            "headers": {
                "Content-Type": "application/json",
                "Access-Control-Allow-Origin": "*",
            },
            "body": json.dumps(
                {
                    "jobId": job_id,
                    "outputS3Key": output_s3_key,
                    "downloadUrl": presigned_url,
                    "stats": stats,
                }
            ),
        }

    except Exception as e:
        logger.exception("Error processing Excel translation")
        error_msg = str(e)
        if job_id:
            update_job_status(job_id, "FAILED", error=error_msg, failedAt=datetime.utcnow().isoformat())
        return {"statusCode": 500, "body": json.dumps({"error": error_msg})}

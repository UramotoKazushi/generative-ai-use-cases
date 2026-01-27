"""
Excel Translation - Prepare Handler for Step Functions

This Lambda:
1. Downloads Excel file from S3
2. Extracts all unique texts
3. Splits into batches for parallel processing
4. Stores batch data in S3
5. Returns batch info for Step Functions Map state
"""

import json
import os
import uuid
from datetime import datetime
from typing import Any

import boto3
from aws_lambda_powertools import Logger
from aws_lambda_powertools.utilities.typing import LambdaContext
from openpyxl import load_workbook
import re

# For .xls support
import xlrd
from openpyxl import Workbook

logger = Logger()

s3_client = boto3.client("s3")
dynamodb_client = boto3.client("dynamodb")

BUCKET_NAME = os.environ.get("BUCKET_NAME", "")
JOB_TABLE_NAME = os.environ.get("JOB_TABLE_NAME", "")
BATCH_SIZE = 100  # Texts per batch


def update_job_status(job_id: str, status: str, **kwargs) -> None:
    """Update job status in DynamoDB"""
    if not JOB_TABLE_NAME or not job_id:
        return

    update_expr = "SET #status = :status"
    expr_names = {"#status": "status"}
    expr_values = {":status": {"S": status}}

    for key, value in kwargs.items():
        update_expr += f", {key} = :{key}"
        if isinstance(value, dict):
            expr_values[f":{key}"] = {"S": json.dumps(value)}
        else:
            expr_values[f":{key}"] = {"S": str(value)}

    try:
        dynamodb_client.update_item(
            TableName=JOB_TABLE_NAME,
            Key={"jobId": {"S": job_id}},
            UpdateExpression=update_expr,
            ExpressionAttributeNames=expr_names,
            ExpressionAttributeValues=expr_values,
        )
    except Exception as e:
        logger.warning(f"Failed to update job status: {e}")


def convert_xls_to_xlsx(xls_path: str, xlsx_path: str) -> None:
    """Convert .xls file to .xlsx format."""
    xls_book = xlrd.open_workbook(xls_path, formatting_info=False)
    xlsx_book = Workbook()

    if "Sheet" in xlsx_book.sheetnames:
        del xlsx_book["Sheet"]

    for sheet_idx in range(xls_book.nsheets):
        xls_sheet = xls_book.sheet_by_index(sheet_idx)
        xlsx_sheet = xlsx_book.create_sheet(title=xls_sheet.name)

        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                cell = xls_sheet.cell(row_idx, col_idx)
                value = cell.value

                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate.xldate_as_datetime(cell.value, xls_book.datemode)
                    except Exception:
                        pass
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    value = bool(cell.value)
                elif cell.ctype == xlrd.XL_CELL_ERROR:
                    value = None

                xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=value)

    xlsx_book.save(xlsx_path)


def is_xls_file(filename: str) -> bool:
    """Check if file is .xls format (not .xlsx)"""
    return filename.lower().endswith(".xls") and not filename.lower().endswith(".xlsx")


def should_skip_text(text: str) -> bool:
    """Check if text should be skipped from translation."""
    text = text.strip()

    if not text:
        return True

    # Numbers only
    if re.match(r'^-?[\d,\s]+\.?\d*%?$', text):
        return True

    # Date patterns
    date_patterns = [
        r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}$',
        r'^\d{1,2}[-/]\d{1,2}[-/]\d{4}$',
        r'^\d{1,2}[-/]\d{1,2}[-/]\d{2}$',
        r'^\d{4}年\d{1,2}月\d{1,2}日$',
    ]
    for pattern in date_patterns:
        if re.match(pattern, text):
            return True

    # Time patterns
    if re.match(r'^\d{1,2}:\d{2}(:\d{2})?(\s*[APap][Mm])?$', text):
        return True

    # URLs
    if re.match(r'^https?://', text, re.IGNORECASE):
        return True

    # Email addresses
    if re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', text):
        return True

    # Phone numbers
    if re.match(r'^[\d\s\-\+\(\)]+$', text) and len(re.sub(r'\D', '', text)) >= 7:
        return True

    # Symbols only
    if not any(c.isalpha() for c in text):
        return True

    # Currency values
    if re.match(r'^[$€£¥₩]\s*[\d,]+\.?\d*$', text):
        return True
    if re.match(r'^[\d,]+\.?\d*\s*[$€£¥₩円]$', text):
        return True

    # ASCII-only short text
    if re.match(r'^[A-Za-z0-9\s\.,;:!?\'"()\-_@#$%&*+=/<>\\|{}\[\]`~]+$', text):
        words = text.split()
        if len(words) <= 2:
            return True
        if re.match(r'^[A-Z][A-Za-z0-9_]+$', text):
            return True
        if re.match(r'^[a-z_][a-z0-9_]*$', text):
            return True

    # File paths
    if re.match(r'^[A-Za-z]:\\', text) or text.startswith('/'):
        return True

    return False


def is_translatable_cell(cell) -> bool:
    """Check if a cell contains translatable text"""
    if cell.value is None:
        return False
    if not isinstance(cell.value, str):
        return False
    if not cell.value.strip():
        return False
    if str(cell.value).startswith("="):
        return False
    if should_skip_text(cell.value):
        return False
    return True


@logger.inject_lambda_context
def lambda_handler(event: dict, context: LambdaContext) -> dict:
    """
    Prepare handler for Step Functions.

    Input:
    {
        "jobId": "uuid",
        "s3Key": "uploads/xxx/file.xlsx",
        "sourceLanguage": "Japanese",
        "targetLanguage": "English"
    }

    Output:
    {
        "jobId": "uuid",
        "workDataKey": "excel-work/jobId/work_data.json",
        "batches": [
            {"batchId": 0, "batchKey": "excel-work/jobId/batch_0.json"},
            {"batchId": 1, "batchKey": "excel-work/jobId/batch_1.json"},
            ...
        ],
        "sourceLanguage": "Japanese",
        "targetLanguage": "English",
        "stats": {...}
    }
    """
    job_id = event.get("jobId")
    s3_key = event.get("s3Key")
    source_lang = event.get("sourceLanguage", "Japanese")
    target_lang = event.get("targetLanguage", "English")

    logger.info(f"Preparing translation job: {job_id}, file: {s3_key}")

    update_job_status(job_id, "PREPARING")

    # Download file
    filename = os.path.basename(s3_key)
    is_xls = is_xls_file(filename)
    unique_id = uuid.uuid4()

    if is_xls:
        tmp_download = f"/tmp/input_{unique_id}.xls"
        tmp_input = f"/tmp/input_{unique_id}_converted.xlsx"
    else:
        tmp_download = f"/tmp/input_{unique_id}.xlsx"
        tmp_input = tmp_download

    s3_client.download_file(BUCKET_NAME, s3_key, tmp_download)
    logger.info(f"Downloaded file from S3: {s3_key}")

    if is_xls:
        convert_xls_to_xlsx(tmp_download, tmp_input)
        logger.info("Converted .xls to .xlsx")

    # Load workbook and extract data
    wb = load_workbook(tmp_input)

    # Collect all cells info and unique texts
    all_cells_info = []  # [(sheet_name, coord, text), ...]
    unique_texts = set()
    total_cell_count = 0

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                total_cell_count += 1
                if is_translatable_cell(cell):
                    text = cell.value
                    all_cells_info.append({
                        "sheet": sheet.title,
                        "coord": cell.coordinate,
                        "text": text
                    })
                    unique_texts.add(text)

    unique_texts_list = list(unique_texts)
    logger.info(f"Found {len(all_cells_info)} translatable cells with {len(unique_texts_list)} unique texts")

    # Split unique texts into batches
    batches = []
    work_prefix = f"excel-work/{job_id}"

    for i in range(0, len(unique_texts_list), BATCH_SIZE):
        batch_texts = unique_texts_list[i:i + BATCH_SIZE]
        batch_id = i // BATCH_SIZE
        batch_key = f"{work_prefix}/batch_{batch_id}.json"

        # Store batch in S3
        s3_client.put_object(
            Bucket=BUCKET_NAME,
            Key=batch_key,
            Body=json.dumps({"texts": batch_texts}),
            ContentType="application/json"
        )

        batches.append({
            "batchId": batch_id,
            "batchKey": batch_key,
            "textCount": len(batch_texts)
        })

    # Store work data (cells info) in S3
    work_data_key = f"{work_prefix}/work_data.json"
    work_data = {
        "s3Key": s3_key,
        "inputPath": tmp_input,
        "cells": all_cells_info,
        "totalCells": total_cell_count,
        "uniqueTexts": len(unique_texts_list),
        "batchCount": len(batches)
    }
    s3_client.put_object(
        Bucket=BUCKET_NAME,
        Key=work_data_key,
        Body=json.dumps(work_data),
        ContentType="application/json"
    )

    # Clean up temp files
    if is_xls and os.path.exists(tmp_download):
        os.remove(tmp_download)
    if os.path.exists(tmp_input):
        os.remove(tmp_input)

    stats = {
        "totalCells": total_cell_count,
        "translatableCells": len(all_cells_info),
        "uniqueTexts": len(unique_texts_list),
        "batchCount": len(batches)
    }

    update_job_status(
        job_id,
        "TRANSLATING",
        progress=json.dumps({
            "phase": "prepared",
            "batches": len(batches),
            "uniqueTexts": len(unique_texts_list),
            "percent": 5
        })
    )

    logger.info(f"Prepared {len(batches)} batches for translation")

    return {
        "jobId": job_id,
        "s3Key": s3_key,
        "workDataKey": work_data_key,
        "batches": batches,
        "totalBatches": len(batches),
        "startTime": datetime.utcnow().isoformat() + "Z",
        "sourceLanguage": source_lang,
        "targetLanguage": target_lang,
        "stats": stats
    }
